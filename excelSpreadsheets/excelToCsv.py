# Changes all excel files to CSV in given path.

import openpyxl, csv, os
from openpyxl.utils import get_column_letter

def convertExcelToCsv():
    for excelFile in os.listdir('.'):
        if excelFile.endswith('.xlsx'):
            print(excelFile)
            wb = openpyxl.load_workbook(excelFile)
            for sheetName in wb.get_sheet_names():
                # Loop through every sheet in the workbook.
                sheet = wb.get_sheet_by_name(sheetName)

                # Create the CSV filename from the Excel filename and sheet title.
                csvFile = open(f'{sheetName}_{excelFile}', 'w', newline='')
                # Create the csv.writer object for this CSV file.
                csvWriter = csv.writer(csvFile)
                
                # Loop through every row in the sheet.
                for rowNum in range(1, sheet.max_row + 1):
                    rowData = []    # append each cell to this list
                    # Loop through each cell in the row.
                    for colNum in range(1, sheet.max_column + 1):
                        # Append each cell's data to rowData.
                        currentCell = get_column_letter(colNum) + str(rowNum)
                        rowData.append(currentCell)

                # Write the rowData list to the CSV file.
                for row in rowData:
                    csvWriter.writerow(row)

                csvFile.close()
            
convertExcelToCsv()